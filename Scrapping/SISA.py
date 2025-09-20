

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os, time, re

URL = "https://sisa.msal.gov.ar/sisa/#sisa"

BASE_DIR = r'C:\Users\herna\Downloads'
XLSX_PATH = os.path.join(BASE_DIR, "comprobaciones_matriculas.xlsx")
TXT_DNIS = os.path.join(BASE_DIR, "dnis.txt")  # DNIs separados por coma

# ---------- Helpers ----------
def append_to_excel(path, row_dict):
    headers = ["Fecha", "DNI", "NombreApellido", "Profesion", "Estado", "Matricula"]
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        ws.append(headers)
        wb.save(path)
    wb = load_workbook(path)
    ws = wb.active
    ws.append([row_dict.get(h, "") for h in headers])
    wb.save(path)
    wb.close()

def extraer_nombre_apellido(titulo_raw: str) -> str:
    if not titulo_raw:
        return ""
    txt = re.sub(r"\s+", " ", titulo_raw).strip()
    m = re.search(r"Ficha personal de\s+(.*?)(,|$)", txt, flags=re.IGNORECASE)
    return m.group(1).strip() if m else txt

def solo_numero_matricula(matricula_raw: str) -> str:
    if not matricula_raw:
        return ""
    m = re.search(r"(\d[\d\.\s]*)", matricula_raw)
    return re.sub(r"[^\d]", "", m.group(1)) if m else re.sub(r"[^\d]", "", matricula_raw)

def leer_dnis_desde_txt(txt_path: str):
    if not os.path.exists(txt_path):
        raise FileNotFoundError(f"No existe el archivo TXT de DNIs: {txt_path}")
    with open(txt_path, "r", encoding="utf-8") as f:
        contenido = f.read()
    crudos = [x.strip() for x in contenido.split(",")]
    dnis, vistos = [], set()
    for x in crudos:
        solo_digitos = re.sub(r"\D", "", x)
        if solo_digitos and solo_digitos not in vistos:
            vistos.add(solo_digitos)
            dnis.append(solo_digitos)
    return dnis

def ir_a_inicio(driver, wait, duro=False):
    """
    Vuelve a la home clickeando el logo. Si falla o duro=True, navega por URL.
    Asegura que el ícono 'Agenda' esté disponible antes de continuar.
    """
    try:
        if not duro:
            logo = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.sisa_logo")))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", logo)
            driver.execute_script("arguments[0].click();", logo)
        else:
            driver.get(URL)
    except Exception:
        driver.get(URL)

    # Esperar a que esté la home (icono Agenda visible/clickable)
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.ico_alcentral.ico_mod_agenda")))

def flujo_refeps_para_dni(driver, wait, dni: str):
    """
    Asume que estamos en la HOME. Hace Agenda->REFEPS->Documento->Buscar->Ficha y devuelve dict con datos.
    """
    # 1) Icono Agenda
    ico_agenda = wait.until(EC.element_to_be_clickable(
        (By.CSS_SELECTOR, "div.ico_alcentral.ico_mod_agenda")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", ico_agenda)
    driver.execute_script("arguments[0].click();", ico_agenda)

    # 2) Tarjeta REFEPS
    ref_eps = wait.until(EC.element_to_be_clickable((
        By.XPATH,
        "("
        "//div[contains(@class,'util_org_hijo') and contains(@class,'botonera_texto')][.//h3[normalize-space()='REFEPS']]"
        "[not(ancestor-or-self::*[@style and contains(translate(@style,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'display: none')])]"
        ")[1]"
    )))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", ref_eps)
    driver.execute_script("arguments[0].click();", ref_eps)

    # 3) Seleccionar “Documento”
    label_doc = wait.until(EC.presence_of_element_located(
        (By.XPATH, "//label[normalize-space(.)='Documento']")))
    driver.execute_script("""
        const lbl = arguments[0];
        const id = lbl.getAttribute('for');
        const inp = id ? document.getElementById(id) : null;
        if (inp) { inp.checked = true; inp.dispatchEvent(new Event('change', {bubbles:true})); }
        else { lbl.click(); }
    """, label_doc)

    # 4) Cargar DNI
    input_doc = wait.until(EC.element_to_be_clickable((
        By.XPATH,
        "//input[@type='text' and contains(@class,'gwt-TextBox') and ("
        "contains(@placeholder,'Documento') or contains(@placeholder,'documento') or contains(@placeholder,'Número'))]"
    )))
    input_doc.clear()
    input_doc.send_keys(dni)

    # 5) Buscar
    buscar_anchor = wait.until(EC.presence_of_element_located((
        By.XPATH,
        "("
        "//a[contains(@class,'gwt-Anchor') and "
        "     .//div[contains(@class,'boton_general') and contains(normalize-space(.),'Buscar')] and "
        "     not(ancestor-or-self::*[@style and contains(translate(@style,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'display: none')])"
        "] )[1]"
    )))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", buscar_anchor)
    driver.execute_script("""
        const e = arguments[0];
        ['mousedown','mouseup','click'].forEach(t =>
          e.dispatchEvent(new MouseEvent(t, {bubbles:true, cancelable:true, view:window}))
        );
    """, buscar_anchor)

    try:
        WebDriverWait(driver, 1.5).until(
            EC.presence_of_element_located((By.XPATH, "//table|//tr[.//td]"))
        )
    except Exception:
        input_doc.send_keys(Keys.ENTER)

    # 6) Ficha (si no hay resultados, no aparece el ícono)
    try:
        ver_ficha = WebDriverWait(driver, 8).until(EC.element_to_be_clickable((
            By.XPATH, "//img[contains(@class,'gwt-Image') and contains(@title,'Ingresar a la ficha')]"
        )))
    except Exception:
        return {
            "Fecha": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "DNI": dni,
            "NombreApellido": "",
            "Profesion": "",
            "Estado": "SIN RESULTADOS",
            "Matricula": "",
        }

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", ver_ficha)
    driver.execute_script("arguments[0].click();", ver_ficha)

    # 7) Extraer datos
    titulo_raw = wait.until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "div.title"))).text.strip()
    nombre_apellido = extraer_nombre_apellido(titulo_raw)

    profesion = wait.until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "div.profesion-Label"))).text.strip()
    estado = wait.until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "div.habilitacion-Label"))).text.strip()
    matricula_raw = wait.until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "div.matricula-Label"))).text.strip()
    matricula_num = solo_numero_matricula(matricula_raw)

    return {
        "Fecha": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "DNI": dni,
        "NombreApellido": nombre_apellido,
        "Profesion": profesion,
        "Estado": estado,
        "Matricula": matricula_num,
    }

# ---------- Main ----------
def main():
    dnis = leer_dnis_desde_txt(TXT_DNIS)
    if not dnis:
        print("No se encontraron DNIs válidos en el TXT.")
        return

    chrome_opts = Options()
    chrome_opts.add_argument("--start-maximized")
    # chrome_opts.add_argument("--headless=new")  # opcional
    driver = webdriver.Chrome(options=chrome_opts)
    wait = WebDriverWait(driver, 45)

    try:
        # entrar una sola vez
        driver.get(URL)
        ir_a_inicio(driver, wait, duro=False)

        for idx, dni in enumerate(dnis, 1):
            print(f"[{idx}/{len(dnis)}] DNI: {dni}")

            # volver SIEMPRE a inicio con el logo antes de cada búsqueda
            ir_a_inicio(driver, wait, duro=False)

            try:
                row = flujo_refeps_para_dni(driver, wait, dni)
            except Exception as e:
                print(f"  ⚠️ Error con DNI {dni}: {e}")
                row = {
                    "Fecha": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    "DNI": dni,
                    "NombreApellido": "",
                    "Profesion": "",
                    "Estado": f"ERROR: {type(e).__name__}",
                    "Matricula": "",
                }

            append_to_excel(XLSX_PATH, row)
            print(f"  ✅ Guardado: {row}")
            time.sleep(1)  # anti-rate-limit

        print(f"\n✔️ Terminado. Excel: {XLSX_PATH}")

    finally:
        try:
            driver.quit()
        except Exception:
            pass

if __name__ == "__main__":
    main()
